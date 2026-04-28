import json
import subprocess
from pathlib import Path
from loguru import logger
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class IntelligenceMaintenance:
    """Clase de mantenimiento para centralizar la lógica de scratch en src."""

    def __init__(self, config_path: Path):
        self.config_path = config_path
        with open(config_path, "r", encoding="utf-8") as f:
            self.config = json.load(f)
        
        self.project_root = config_path.parent.parent

    def apply_excel_links(self, artifact_id: str):
        """Aplica los enlaces de referencia al Excel especificado en el mapa."""
        if artifact_id not in self.config["data_lineage"]:
            logger.error(f"Artifact ID {artifact_id} no encontrado en el mapa.")
            return

        data_info = self.config["data_lineage"][artifact_id]
        excel_path = self.project_root / data_info["excel_path"]
        links = data_info["variable_mapping"]

        if not excel_path.exists():
            logger.error(f"No se encontró el Excel en: {excel_path}")
            return

        logger.info(f"Aplicando enlaces a: {excel_path.name}")
        wb = openpyxl.load_workbook(excel_path)
        
        if "Diccionario" not in wb.sheetnames:
            logger.warning("No se encontró la pestaña 'Diccionario'.")
            return

        ws = wb["Diccionario"]
        
        # Estilo de encabezado
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws.cell(row=1, column=6, value="Link de Referencia")
        ws.cell(row=1, column=6).fill = header_fill
        ws.cell(row=1, column=6).font = header_font
        ws.cell(row=1, column=6).alignment = Alignment(horizontal="center", vertical="center")

        for row in range(2, ws.max_row + 1):
            var_name = ws.cell(row=row, column=1).value
            if var_name in links:
                cell = ws.cell(row=row, column=6, value=links[var_name])
                cell.font = Font(color="0563C1", underline="single")
            else:
                ws.cell(row=row, column=6, value="N/A (Transformada)")

        ws.column_dimensions['F'].width = 60
        wb.save(excel_path)
        logger.success(f"Enlaces aplicados exitosamente en {excel_path.name}")

    def launch_pending_ocr(self):
        """Lanza el motor pesado (ocrmypdf) para los archivos pendientes."""
        pipeline = self.config["ocr_pipeline"]
        registry = pipeline["registry"]
        
        launched = 0
        for entry in registry:
            if entry["status"] == "PENDING":
                source_path = self.project_root / entry["source"]
                if source_path.exists():
                    output_pdf = str(source_path).replace(".pdf", pipeline["parameters"]["output_suffix"])
                    force_flag = "--force-ocr" if pipeline["parameters"]["force_ocr"] else ""
                    
                    cmd = f"ocrmypdf {force_flag} '{source_path}' '{output_pdf}' > '{output_pdf}.log' 2>&1 &"
                    logger.info(f"🚀 Lanzando OCR para: {source_path.name}")
                    subprocess.Popen(cmd, shell=True)
                    launched += 1
                else:
                    logger.warning(f"No se encontró el origen: {entry['source']}")

        if launched == 0:
            logger.info("No hay OCRs pendientes en el registro.")
        return launched

    def reset_rag(self):
        """Borra la colección RAG para una reconstrucción limpia."""
        from ecs_quantitative.ingestion.rag import MarkdownRAG
        collection = self.config["rag_config"]["collection_name"]
        
        logger.warning(f"Reseteando colección RAG: {collection}")
        rag = MarkdownRAG(collection_name=collection)
        # Asumiendo que existe un método para borrar o simplemente reinicializar
        # En ecs_quantitative, esto suele hacerse borrando el persist_directory o vía API
        logger.info("RAG Reset solicitado (Simulado - Requiere integración con ChromaDB/Qdrant)")

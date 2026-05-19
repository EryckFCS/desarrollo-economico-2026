import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# Ensure directories exist
os.makedirs("docs/vaults/u2-aa-01-expanded-hdi/data", exist_ok=True)
os.makedirs("docs/vaults/u2-aa-01-expanded-hdi/assets", exist_ok=True)
os.makedirs("docs/vaults/u2-aa-01-expanded-hdi/logs", exist_ok=True)

log_path = "docs/vaults/u2-aa-01-expanded-hdi/logs/calculate_indices.log"

def log_message(message):
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(message + "\n")
    print(message)

# Clear log
with open(log_path, "w", encoding="utf-8") as f:
    f.write("=== LOG DE CÁLCULO DE ÍNDICE AMPLIADO DE DESARROLLO HUMANO (PIVOTE A DESEMPLEO) ===\n")

log_message("Inicializando variables y vectores de datos históricos para Ecuador (2013-2022)...")

# 1. Raw Data Vectors
# 1. Load Raw Data Dynamically from Ecuador.csv (HDR 2023/2024 database)
csv_raw_path = "docs/vaults/u2-aa-01-expanded-hdi/data/raw/Ecuador.csv"
df_raw = pd.read_csv(csv_raw_path, index_col="key")

years = np.arange(2013, 2023)
official_hdi = np.array([float(df_raw.loc[f"Human Development Index (value) ({y})", "value"]) for y in years])
health_raw = np.array([float(df_raw.loc[f"Life Expectancy at Birth (years) ({y})", "value"]) for y in years])
expected_raw = np.array([float(df_raw.loc[f"Expected Years of Schooling (years) ({y})", "value"]) for y in years])
mean_raw = np.array([float(df_raw.loc[f"Mean Years of Schooling (years) ({y})", "value"]) for y in years])
gni_raw = np.array([float(df_raw.loc[f"Gross National Income Per Capita (2021 PPP$) ({y})", "value"]) for y in years])

# Calculate standard component indices dynamically from raw indicators
health_idx = (health_raw - 20.0) / 65.0
expected_idx = expected_raw / 18.0
mean_idx = mean_raw / 15.0
edu_idx = (expected_idx + mean_idx) / 2.0
income_idx = (np.log(gni_raw) - np.log(100.0)) / (np.log(75000.0) - np.log(100.0))

# Unemployment rates (%) - Source: World Bank / ILO modeled estimates
unemployment_rates = np.array([3.08, 3.48, 3.62, 4.60, 3.84, 3.53, 3.81, 6.13, 4.55, 3.76])

log_message("Vectores cargados exitosamente.")
log_message("Aplicando fórmula de normalización negativa para la Tasa de Desempleo (Max=15%, Min=0%)...")

# 2. Normalize Additional Indicator (Employment Security Index)
# Since unemployment is a negative indicator: Index = (Max - Value) / (Max - Min)
max_unemp = 15.0
min_unemp = 0.0
employment_idx = (max_unemp - unemployment_rates) / (max_unemp - min_unemp)

log_message("Calculando Índice Ampliado de Desarrollo Humano (Media Geométrica de 4 dimensiones)...")

# 3. Calculate Expanded HDI (geometric mean of 4 components)
expanded_hdi = (health_idx * edu_idx * income_idx * employment_idx) ** 0.25

# 4. Construct DataFrame
df = pd.DataFrame({
    "Año": years,
    "Esperanza_Vida_Raw": health_raw,
    "Índice_Salud": health_idx,
    "Años_Esperados_Raw": expected_raw,
    "Años_Promedio_Raw": mean_raw,
    "Índice_Educación": edu_idx,
    "INB_per_Capita_Raw": gni_raw,
    "Índice_Ingresos": income_idx,
    "Tasa_Desempleo_Porc": unemployment_rates,
    "Índice_Empleo": employment_idx,
    "IDH_Oficial": official_hdi,
    "IDH_Ampliado": expanded_hdi
})

# Save to CSV
csv_path = "docs/vaults/u2-aa-01-expanded-hdi/data/ecuador_hdi_expanded.csv"
df.to_csv(csv_path, index=False, encoding="utf-8")
log_message(f"Archivo de datos intermedio guardado en CSV: {csv_path}")

# 5. Generate Academic Comparison Chart (For PDF and Word Reports)
log_message("Generando gráfico comparativo de alta definición para reportes...")
plt.figure(figsize=(10, 6))
sns.set_theme(style="whitegrid")

# Plot Lines
plt.plot(years, official_hdi, label="IDH Oficial (PNUD)", color="#1F497D", marker="o", linewidth=2.5, markersize=8)
plt.plot(years, expanded_hdi, label="IDH Ampliado (IDHA con Empleo)", color="#E67E22", marker="s", linewidth=2.5, markersize=8)

# Formatting
plt.title("Ecuador: Comparación entre el IDH Oficial y el Índice Ampliado con Desempleo (2013-2022)", fontsize=14, fontweight="bold", pad=15)
plt.xlabel("Año", fontsize=12, fontweight="semibold", labelpad=10)
plt.ylabel("Valor del Índice", fontsize=12, fontweight="semibold", labelpad=10)
plt.xticks(years)
plt.ylim(0.55, 0.82)

# Highlight Covid drop in 2020
plt.axvspan(2019.8, 2020.2, color="#FADBD8", alpha=0.4, label="Recesión y Desempleo Pandémico (COVID-19)")
# Highlight 2016 Recession
plt.axvspan(2015.8, 2016.2, color="#FCF3CF", alpha=0.4, label="Recesión Macroeconómica 2016")

# Annotate critical gaps
plt.annotate(
    f"Brecha COVID-19 (2020):\nΔ = {official_hdi[7] - expanded_hdi[7]:.3f} (-5.5%)",
    xy=(2020, expanded_hdi[7]),
    xytext=(2016.5, 0.61),
    arrowprops=dict(facecolor='black', shrink=0.08, width=1.5, headwidth=8),
    fontsize=10,
    fontweight="bold",
    bbox=dict(boxstyle="round,pad=0.5", fc="#FFF2CC", ec="#D6B656", lw=1.5)
)

# Legend and Grid
plt.legend(loc="lower left", fontsize=10, frameon=True, facecolor="white", edgecolor="lightgray")
plt.tight_layout()

chart_path = "docs/vaults/u2-aa-01-expanded-hdi/assets/hdi_comparison.png"
plt.savefig(chart_path, dpi=300)
plt.close()
log_message(f"Gráfico guardado en: {chart_path}")

# 6. Create Styled Excel Sheet with Live Formulas using openpyxl
log_message("Creando libro de Excel con fórmulas vivas y formato institucional...")
wb = Workbook()

# --- HOJA 1: Calculos_IDHA ---
ws = wb.active
ws.title = "Calculos_IDHA"
ws.views.sheetView[0].showGridLines = True

# Styles
navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
bold_font = Font(name="Calibri", size=11, bold=True)
regular_font = Font(name="Calibri", size=11)
title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center")
right_alignment = Alignment(horizontal="right", vertical="center")

thin_side = Side(border_style="thin", color="D3D3D3")
double_side = Side(border_style="double", color="000000")
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_total = Border(top=thin_side, bottom=double_side)

# Title block
ws.merge_cells("A1:N1")
ws["A1"] = "Índice Ampliado de Desarrollo Humano (IDHA) - Ecuador (2013-2022)"
ws["A1"].font = title_font
ws["A1"].alignment = left_alignment
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:N2")
ws["A2"] = "Materia: Desarrollo Económico | Tarea: Unidad 2 AA1 | Autores: Condoy Seraquive Erick Fabricio & Bustamante Dayana"
ws["A2"].font = Font(name="Calibri", size=10, italic=True)
ws["A2"].alignment = left_alignment

headers = [
    "Año", 
    "Esp. Vida (Años) [Crudo]", 
    "Índice Salud (I_Salud)", 
    "Años Esp. Escolaridad [Crudo]", 
    "Índice Años Esp.", 
    "Años Prom. Escolaridad [Crudo]", 
    "Índice Años Prom.", 
    "Índice Educación (I_Edu)", 
    "INB per cápita PPA (USD) [Crudo]", 
    "Índice Ingresos (I_Ing)", 
    "Tasa Desempleo (%) [Crudo]", 
    "Índice Empleo (I_Emp)", 
    "IDH Oficial", 
    "IDH Ampliado"
]

ws.row_dimensions[4].height = 25
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.fill = navy_fill
    cell.font = white_font
    cell.alignment = center_alignment
    cell.border = border_all

# Populate data and formulas
for i in range(len(years)):
    row_num = 5 + i
    ws.row_dimensions[row_num].height = 20
    
    # Year (string category)
    ws.cell(row=row_num, column=1, value=str(int(years[i]))).alignment = center_alignment
    
    # 2. Esperanza de Vida (Crudo)
    ws.cell(row=row_num, column=2, value=float(health_raw[i])).number_format = "0.000"
    
    # 3. Índice de Salud (Normalizado con fórmula viva)
    ws.cell(row=row_num, column=3, value=f"=(B{row_num}-20)/65").number_format = "0.000"
    
    # 4. Años Esperados de Escolaridad (Crudo)
    ws.cell(row=row_num, column=4, value=float(expected_raw[i])).number_format = "0.000"
    
    # 5. Índice Años Esperados (Normalizado con fórmula viva)
    ws.cell(row=row_num, column=5, value=f"=D{row_num}/18").number_format = "0.000"
    
    # 6. Años Promedio de Escolaridad (Crudo)
    ws.cell(row=row_num, column=6, value=float(mean_raw[i])).number_format = "0.000"
    
    # 7. Índice Años Promedio (Normalizado con fórmula viva)
    ws.cell(row=row_num, column=7, value=f"=F{row_num}/15").number_format = "0.000"
    
    # 8. Índice de Educación (Fórmula promedio simple de componentes)
    ws.cell(row=row_num, column=8, value=f"=(E{row_num}+G{row_num})/2").number_format = "0.000"
    
    # 9. INB per Cápita PPA (Crudo)
    ws.cell(row=row_num, column=9, value=float(gni_raw[i])).number_format = "$#,##0.00"
    
    # 10. Índice de Ingresos (Normalizado con fórmula logarítmica viva)
    ws.cell(row=row_num, column=10, value=f"=(LN(I{row_num})-LN(100))/(LN(75000)-LN(100))").number_format = "0.000"
    
    # 11. Tasa de Desempleo (Cruda, convertida a porcentaje decimal)
    ws.cell(row=row_num, column=11, value=float(unemployment_rates[i]) / 100.0).number_format = "0.00%"
    
    # 12. Índice de Empleo (Fórmula normalización negativa)
    ws.cell(row=row_num, column=12, value=f"=(0.15-K{row_num})/0.15").number_format = "0.000"
    
    # 13. IDH Oficial (Fórmula media geométrica 3 dimensiones)
    ws.cell(row=row_num, column=13, value=f"=(C{row_num}*H{row_num}*J{row_num})^(1/3)").number_format = "0.000"
    
    # 14. IDH Ampliado (Fórmula media geométrica 4 dimensiones)
    ws.cell(row=row_num, column=14, value=f"=(C{row_num}*H{row_num}*J{row_num}*L{row_num})^(1/4)").number_format = "0.000"
    
    for col_idx in range(1, 15):
        c = ws.cell(row=row_num, column=col_idx)
        c.font = regular_font
        c.border = border_all
        if col_idx > 1:
            c.alignment = right_alignment
        if i % 2 == 1:
            c.fill = light_gray_fill

# Add statistics row (Average)
avg_row = 5 + len(years)
ws.row_dimensions[avg_row].height = 22
ws.cell(row=avg_row, column=1, value="Promedio").font = bold_font
ws.cell(row=avg_row, column=1).alignment = center_alignment
ws.cell(row=avg_row, column=1).border = border_total

for col_idx in range(2, 15):
    col_letter = get_column_letter(col_idx)
    cell = ws.cell(row=avg_row, column=col_idx, value=f"=AVERAGE({col_letter}5:{col_letter}{avg_row-1})")
    cell.font = bold_font
    cell.alignment = right_alignment
    cell.border = border_total
    
    if col_idx in [2, 4, 6]:
        cell.number_format = "0.000"
    elif col_idx == 9:
        cell.number_format = "$#,##0.00"
    elif col_idx == 11:
        cell.number_format = "0.00%"
    else:
        cell.number_format = "0.000"

# Adjust columns width dynamically
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
         if cell.value:
             max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 14)


# --- HOJA 2: Figure (GENERATED NATIVELY IN EXCEL WITH CUSTOM STYLING) ---
log_message("Creando hoja 'Figure' y generando gráfica nativa de Excel altamente estilizada...")
ws_fig = wb.create_sheet(title="Figure")
ws_fig.views.sheetView[0].showGridLines = True

# Title block in Figure sheet
ws_fig.row_dimensions[1].height = 25
ws_fig.cell(row=1, column=1, value="Comparación de Resultados (Gráfico Nativo de Excel)").font = title_font
ws_fig.cell(row=1, column=1).alignment = left_alignment

# Create LineChart
chart = LineChart()
chart.title = "Ecuador: Comparación entre el IDH Oficial y el Índice Ampliado (2013-2022)"
chart.style = 13  # Modern clean line chart style
chart.y_axis.title = "Valor del Índice"
chart.x_axis.title = "Año"
chart.width = 18
chart.height = 12

# Explicitly force axes to render and not delete
chart.x_axis.delete = False
chart.y_axis.delete = False

# References for data and categories in sheet "Calculos_IDHA"
# Data: Columns M and N (IDH Oficial in 13, IDH Ampliado in 14) from row 4 (header) to row 14 (end of series)
data_ref = Reference(ws, min_col=13, min_row=4, max_col=14, max_row=14)
# Categories: Column A (Years) from row 5 to 14
cats_ref = Reference(ws, min_col=1, min_row=5, max_row=14)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

# --- ADVANCED PREMIUM STYLING ---
s1 = chart.series[0]
s1.graphicalProperties.line.solidFill = "1F497D" # Navy Blue
s1.graphicalProperties.line.width = 38100       # 3 pt thickness
s1.marker.symbol = "circle"
s1.marker.size = 7
s1.marker.graphicalProperties.solidFill = "1F497D"
s1.marker.graphicalProperties.line.solidFill = "1F497D"
s1.smooth = True

s2 = chart.series[1]
s2.graphicalProperties.line.solidFill = "E67E22" # Accent Orange
s2.graphicalProperties.line.width = 38100       # 3 pt thickness
s2.marker.symbol = "square"
s2.marker.size = 7
s2.marker.graphicalProperties.solidFill = "E67E22"
s2.marker.graphicalProperties.line.solidFill = "E67E22"
s2.smooth = True

# Add chart to sheet
ws_fig.add_chart(chart, "A3")


# --- HOJA 3: Dictionary ---
log_message("Creando hoja 'Dictionary' y populando metadatos de variables...")
ws_dict = wb.create_sheet(title="Dictionary")
ws_dict.views.sheetView[0].showGridLines = True

# Title block in Dictionary sheet
ws_dict.merge_cells("A1:F1")
ws_dict["A1"] = "Diccionario de Variables y Metadatos de Trazabilidad"
ws_dict["A1"].font = title_font
ws_dict["A1"].alignment = left_alignment
ws_dict.row_dimensions[1].height = 30

ws_dict.merge_cells("A2:F2")
ws_dict["A2"] = "Detalle técnico, origen institucional, fórmulas y enlaces de descarga de las variables del proyecto."
ws_dict["A2"].font = Font(name="Calibri", size=10, italic=True)
ws_dict["A2"].alignment = left_alignment

dict_headers = ["Variable", "Columna Excel", "Definición Conceptual", "Unidad de Medida", "Fuente de Origen", "Enlace de Descarga (Link)"]
ws_dict.row_dimensions[4].height = 25
for col_idx, header in enumerate(dict_headers, 1):
    cell = ws_dict.cell(row=4, column=col_idx, value=header)
    cell.fill = navy_fill
    cell.font = white_font
    cell.alignment = center_alignment
    cell.border = border_all

dict_data = [
    {
        "variable": "Año (t)",
        "column": "A",
        "definition": "Periodo anual cronológico continuo analizado en la serie temporal.",
        "unit": "Año (Texto para compatibilidad de gráficos)",
        "source": "Definición metodológica de la tarea.",
        "link": "N/A"
    },
    {
        "variable": "Esperanza de Vida al Nacer (Años) [Crudo]",
        "column": "B",
        "definition": "Número promedio de años que se espera que viva un recién nacido si se mantienen los patrones de mortalidad vigentes.",
        "unit": "Años",
        "source": "PNUD (Human Development Reports)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Índice de Salud (I_Salud) [Normalizado]",
        "column": "C",
        "definition": "Subíndice oficial de la dimensión de salud, normalizado mediante la fórmula =(B5 - 20) / 65.",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Cálculo mediante fórmula oficial)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Años Esperados de Escolaridad (Años) [Crudo]",
        "column": "D",
        "definition": "Número de años que se espera que un niño en edad de ingresar a la escuela asista a ella.",
        "unit": "Años",
        "source": "PNUD (Human Development Reports)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Índice de Años Esperados [Normalizado]",
        "column": "E",
        "definition": "Subíndice de educación por años esperados, normalizado mediante la fórmula =D5 / 18.",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Cálculo mediante fórmula oficial)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Años Promedio de Escolaridad (Años) [Crudo]",
        "column": "F",
        "definition": "Número promedio de años de educación formal completados por la población de 25 años o más.",
        "unit": "Años",
        "source": "PNUD (Human Development Reports)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Índice de Años Promedio [Normalizado]",
        "column": "G",
        "definition": "Subíndice de educación por años promedio, normalizado mediante la fórmula =F5 / 15.",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Cálculo mediante fórmula oficial)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Índice de Educación (I_Edu) [Combinado]",
        "column": "H",
        "definition": "Subíndice oficial de la dimensión educativa, calculado como el promedio simple del Índice de Años Esperados y el de Años Promedio: =(E5+G5)/2.",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Cálculo mediante fórmula oficial)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Ingreso Nacional Bruto (INB) per cápita PPA (USD) [Crudo]",
        "column": "I",
        "definition": "Suma del valor agregado por todos los productores residentes en la economía, más los flujos netos de ingreso primario del exterior, expresado en dólares internacionales constantes mediante la Paridad de Poder Adquisitivo.",
        "unit": "Dólares Internacionales PPA",
        "source": "PNUD (Human Development Reports)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Índice de Ingresos (I_Ing) [Normalizado]",
        "column": "J",
        "definition": "Subíndice oficial de estándar de vida, normalizado mediante la transformación logarítmica =(LN(I5)-LN(100))/(LN(75000)-LN(100)).",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Cálculo mediante fórmula oficial)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "Tasa de Desempleo (%) [Crudo]",
        "column": "K",
        "definition": "Porcentaje de la población económicamente activa que no tiene trabajo pero busca activamente empleo.",
        "unit": "Porcentaje (%)",
        "source": "Banco Mundial / OIT",
        "link": "https://databank.worldbank.org/source/world-development-indicators"
    },
    {
        "variable": "Índice de Empleo (I_Emp) [Normalizado]",
        "column": "L",
        "definition": "Indicador de seguridad económica en el empleo, normalizado inversamente de la tasa de desempleo mediante la fórmula =(0.15-K5)/0.15.",
        "unit": "Índice (0 a 1)",
        "source": "Elaboración propia con base en el Enfoque de Capacidades de Amartya Sen.",
        "link": "Elaboración propia"
    },
    {
        "variable": "IDH Oficial",
        "column": "M",
        "definition": "Índice de Desarrollo Humano oficial tradicional, calculado como la media geométrica de las dimensiones estándar: =(C5*H5*J5)^(1/3).",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Cálculo mediante fórmula oficial)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "IDH Ampliado (IDHA)",
        "column": "N",
        "definition": "Índice de Desarrollo Humano Ampliado propuesto, calculado como la media geométrica de las cuatro dimensiones con igual ponderación: =(C5*H5*J5*L5)^(1/4).",
        "unit": "Índice (0 a 1)",
        "source": "Elaboración propia",
        "link": "Elaboración propia"
    }
]

for row_idx, data in enumerate(dict_data, 5):
    ws_dict.row_dimensions[row_idx].height = 32
    
    ws_dict.cell(row=row_idx, column=1, value=data["variable"]).font = bold_font
    ws_dict.cell(row=row_idx, column=1).alignment = left_alignment
    ws_dict.cell(row=row_idx, column=1).border = border_all
    
    ws_dict.cell(row=row_idx, column=2, value=data["column"]).font = regular_font
    ws_dict.cell(row=row_idx, column=2).alignment = center_alignment
    ws_dict.cell(row=row_idx, column=2).border = border_all
    
    c_def = ws_dict.cell(row=row_idx, column=3, value=data["definition"])
    c_def.font = regular_font
    c_def.alignment = left_alignment
    c_def.border = border_all
    c_def.alignment = Alignment(wrap_text=True, vertical="center")
    
    ws_dict.cell(row=row_idx, column=4, value=data["unit"]).font = regular_font
    ws_dict.cell(row=row_idx, column=4).alignment = left_alignment
    ws_dict.cell(row=row_idx, column=4).border = border_all
    
    c_src = ws_dict.cell(row=row_idx, column=5, value=data["source"])
    c_src.font = regular_font
    c_src.alignment = left_alignment
    c_src.border = border_all
    c_src.alignment = Alignment(wrap_text=True, vertical="center")
    
    cell_link = ws_dict.cell(row=row_idx, column=6, value=data["link"])
    cell_link.border = border_all
    if data["link"].startswith("http"):
        cell_link.hyperlink = data["link"]
        cell_link.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        cell_link.alignment = left_alignment
    else:
        cell_link.font = regular_font
        cell_link.alignment = left_alignment
        
    if row_idx % 2 == 1:
        for col_idx in range(1, 7):
            ws_dict.cell(row=row_idx, column=col_idx).fill = light_gray_fill

# Adjust column widths for Dictionary
ws_dict.column_dimensions["A"].width = 25
ws_dict.column_dimensions["B"].width = 15
ws_dict.column_dimensions["C"].width = 45
ws_dict.column_dimensions["D"].width = 25
ws_dict.column_dimensions["E"].width = 30
ws_dict.column_dimensions["F"].width = 35

# Save final structured workbook
excel_path = "docs/vaults/u2-aa-01-expanded-hdi/data/ecuador_hdi_expanded.xlsx"
wb.save(excel_path)
log_message(f"Libro de Excel maestro con hojas 'Calculos_IDHA', 'Figure' y 'Dictionary' guardado en: {excel_path}")

log_message("=== CÁLCULO COMPLETADO CON ÉXITO ===")

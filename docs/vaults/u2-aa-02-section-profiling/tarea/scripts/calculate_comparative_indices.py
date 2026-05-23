import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# Setup directories
base_dir = "/home/erick-fcs/Documentos/universidad/07_Ciclo/septimo_ciclo/economic_development/docs/vaults/u2-aa-02-section-profiling/tarea"
os.makedirs(f"{base_dir}/data/processed", exist_ok=True)
os.makedirs(f"{base_dir}/assets", exist_ok=True)
os.makedirs(f"{base_dir}/logs", exist_ok=True)

log_path = f"{base_dir}/logs/calculate_comparative_indices.log"

def log_message(message):
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(message + "\n")
    print(message)

# Clear log
with open(log_path, "w", encoding="utf-8") as f:
    f.write("=== LOG DE CÁLCULO COMPARATIVO COMPLETO (IDH, IHDI, PHDI, GII, IDHA) ===\n")

log_message("Inicializando extracción de datos de PNUD...")

ecuador_raw_path = f"{base_dir}/data/raw/Ecuador.csv"
switzerland_raw_path = f"{base_dir}/data/raw/Switzerland.csv"

df_ecu_raw = pd.read_csv(ecuador_raw_path, index_col="key")
df_che_raw = pd.read_csv(switzerland_raw_path, index_col="key")

years = np.arange(2000, 2021)

def safe_extract(df, key_prefix, years):
    values = []
    for y in years:
        key = f"{key_prefix} ({y})"
        if key in df.index:
            try:
                val = df.loc[key, "value"]
                # In case of multiple entries, take the first
                if isinstance(val, pd.Series):
                    val = val.iloc[0]
                values.append(float(val))
            except Exception as e:
                values.append(np.nan)
        else:
            values.append(np.nan)
    return np.array(values)

# 1. Extract PNUD indicators for Ecuador
ecu_official_hdi = safe_extract(df_ecu_raw, "Human Development Index (value)", years)
ecu_health_raw = safe_extract(df_ecu_raw, "Life Expectancy at Birth (years)", years)
ecu_expected_raw = safe_extract(df_ecu_raw, "Expected Years of Schooling (years)", years)
ecu_mean_raw = safe_extract(df_ecu_raw, "Mean Years of Schooling (years)", years)
ecu_gni_raw = safe_extract(df_ecu_raw, "Gross National Income Per Capita (2021 PPP$)", years)
ecu_ihdi_raw = safe_extract(df_ecu_raw, "Inequality-adjusted Human Development Index (value)", years)
ecu_phdi_raw = safe_extract(df_ecu_raw, "Planetary pressures-adjusted Human Development Index (value)", years)
ecu_gii_raw = safe_extract(df_ecu_raw, "Gender Inequality Index (value)", years)

# 2. Extract PNUD indicators for Switzerland
che_official_hdi = safe_extract(df_che_raw, "Human Development Index (value)", years)
che_health_raw = safe_extract(df_che_raw, "Life Expectancy at Birth (years)", years)
che_expected_raw = safe_extract(df_che_raw, "Expected Years of Schooling (years)", years)
che_mean_raw = safe_extract(df_che_raw, "Mean Years of Schooling (years)", years)
che_gni_raw = safe_extract(df_che_raw, "Gross National Income Per Capita (2021 PPP$)", years)
che_ihdi_raw = safe_extract(df_che_raw, "Inequality-adjusted Human Development Index (value)", years)
che_phdi_raw = safe_extract(df_che_raw, "Planetary pressures-adjusted Human Development Index (value)", years)
che_gii_raw = safe_extract(df_che_raw, "Gender Inequality Index (value)", years)

# 3. Load Unemployment Rates (World Bank / ILO modeled estimates 2000-2020)
ecu_unemployment = np.array([4.80, 4.25, 4.84, 5.66, 5.00, 3.78, 3.55, 3.14, 3.92, 4.61, 4.09, 3.46, 3.24, 3.08, 3.48, 3.62, 4.60, 3.84, 3.53, 3.81, 6.13])
che_unemployment = np.array([2.66, 2.49, 2.92, 4.12, 4.41, 4.42, 3.73, 3.16, 3.03, 3.99, 4.36, 4.09, 4.25, 4.43, 4.45, 4.54, 4.67, 4.71, 4.66, 4.35, 4.81])

# 4. Perform calculations in Python for validation and CSV output
def compute_hdi_components(health, expected, mean, gni, unemp):
    i_health = (health - 20.0) / 65.0
    i_expected = expected / 18.0
    i_mean = mean / 15.0
    i_edu = (i_expected + i_mean) / 2.0
    # Apply standard PNUD cap of 75000
    gni_capped = np.minimum(gni, 75000.0)
    i_income = (np.log(gni_capped) - np.log(100.0)) / (np.log(75000.0) - np.log(100.0))
    i_emp = (15.0 - unemp) / 15.0
    idh_ofic_calc = (i_health * i_edu * i_income) ** (1.0/3.0)
    idha = (i_health * i_edu * i_income * i_emp) ** 0.25
    return i_health, i_edu, i_income, i_emp, idh_ofic_calc, idha

ecu_h, ecu_e, ecu_i, ecu_emp, ecu_idh_calc, ecu_idha = compute_hdi_components(ecu_health_raw, ecu_expected_raw, ecu_mean_raw, ecu_gni_raw, ecu_unemployment)
che_h, che_e, che_i, che_emp, che_idh_calc, che_idha = compute_hdi_components(che_health_raw, che_expected_raw, che_mean_raw, che_gni_raw, che_unemployment)

# Create DataFrames with all requested indicators
df_ecu = pd.DataFrame({
    "Año": years,
    "Esperanza_Vida_Raw": ecu_health_raw,
    "Índice_Salud": ecu_h,
    "Años_Esperados_Raw": ecu_expected_raw,
    "Años_Promedio_Raw": ecu_mean_raw,
    "Índice_Educación": ecu_e,
    "INB_per_Capita_Raw": ecu_gni_raw,
    "Índice_Ingresos": ecu_i,
    "Tasa_Desempleo_Porc": ecu_unemployment,
    "Índice_Empleo": ecu_emp,
    "IDH_Oficial_Raw": ecu_official_hdi,
    "IDH_Oficial_Calc": ecu_idh_calc,
    "IHDI_Raw": ecu_ihdi_raw,
    "PHDI_Raw": ecu_phdi_raw,
    "GII_Raw": ecu_gii_raw,
    "IDH_Ampliado": ecu_idha
})

df_che = pd.DataFrame({
    "Año": years,
    "Esperanza_Vida_Raw": che_health_raw,
    "Índice_Salud": che_h,
    "Años_Esperados_Raw": che_expected_raw,
    "Años_Promedio_Raw": che_mean_raw,
    "Índice_Educación": che_e,
    "INB_per_Capita_Raw": che_gni_raw,
    "Índice_Ingresos": che_i,
    "Tasa_Desempleo_Porc": che_unemployment,
    "Índice_Empleo": che_emp,
    "IDH_Oficial_Raw": che_official_hdi,
    "IDH_Oficial_Calc": che_idh_calc,
    "IHDI_Raw": che_ihdi_raw,
    "PHDI_Raw": che_phdi_raw,
    "GII_Raw": che_gii_raw,
    "IDH_Ampliado": che_idha
})

df_comp = pd.DataFrame({
    "Año": years,
    "Ecuador_IDH": ecu_official_hdi,
    "Suiza_IDH": che_official_hdi,
    "Ecuador_IHDI": ecu_ihdi_raw,
    "Suiza_IHDI": che_ihdi_raw,
    "Ecuador_PHDI": ecu_phdi_raw,
    "Suiza_PHDI": che_phdi_raw,
    "Ecuador_GII": ecu_gii_raw,
    "Suiza_GII": che_gii_raw,
    "Ecuador_IDHA": ecu_idha,
    "Suiza_IDHA": che_idha
})

# Save to CSV
df_ecu.to_csv(f"{base_dir}/data/processed/ecuador_hdi_expanded_20y.csv", index=False, encoding="utf-8")
df_che.to_csv(f"{base_dir}/data/processed/switzerland_hdi_expanded_20y.csv", index=False, encoding="utf-8")
df_comp.to_csv(f"{base_dir}/data/processed/comparative_hdi_expanded_20y.csv", index=False, encoding="utf-8")
log_message("CSVs guardados exitosamente en data/processed/.")

# 5. Generate high-definition comparison charts using Matplotlib / Seaborn
sns.set_theme(style="whitegrid")

# Chart 1: IDH Comparison
plt.figure(figsize=(10, 6))
plt.plot(years, che_official_hdi, label="Suiza (IDH)", color="#2C3E50", marker="o", linewidth=2.5)
plt.plot(years, ecu_official_hdi, label="Ecuador (IDH)", color="#3498DB", marker="s", linewidth=2.5)
plt.title("Comparativo Internacional: Índice de Desarrollo Humano (IDH) (2000-2020)", fontsize=12, fontweight="bold", pad=15)
plt.xlabel("Año", fontsize=10, fontweight="semibold")
plt.ylabel("IDH", fontsize=10, fontweight="semibold")
plt.xticks(years, rotation=45)
plt.ylim(0.60, 1.00)
plt.legend(loc="lower right", frameon=True, facecolor="white")
plt.tight_layout()
plt.savefig(f"{base_dir}/assets/hdi_comparison.png", dpi=300)
plt.close()

# Chart 2: IHDI Comparison (Desigualdad, available 2010-2020)
plt.figure(figsize=(10, 6))
valid_idx = ~np.isnan(ecu_ihdi_raw)
plt.plot(years[valid_idx], che_ihdi_raw[valid_idx], label="Suiza (IDH Ajustado por Desigualdad - IHDI)", color="#27AE60", marker="o", linewidth=2.5)
plt.plot(years[valid_idx], ecu_ihdi_raw[valid_idx], label="Ecuador (IDH Ajustado por Desigualdad - IHDI)", color="#F1C40F", marker="s", linewidth=2.5)
plt.title("Comparativo Internacional: IDH Ajustado por Desigualdad (IHDI) (2000-2020)", fontsize=12, fontweight="bold", pad=15)
plt.xlabel("Año", fontsize=10, fontweight="semibold")
plt.ylabel("IHDI", fontsize=10, fontweight="semibold")
plt.xticks(years[valid_idx], rotation=45)
plt.ylim(0.50, 1.00)
plt.legend(loc="lower right", frameon=True, facecolor="white")
plt.tight_layout()
plt.savefig(f"{base_dir}/assets/ihdi_comparison.png", dpi=300)
plt.close()

# Chart 3: PHDI Comparison (Presiones Planetarias)
plt.figure(figsize=(10, 6))
plt.plot(years, che_phdi_raw, label="Suiza (IDH Ajustado por Presiones Planetarias - PHDI)", color="#8E44AD", marker="o", linewidth=2.5)
plt.plot(years, ecu_phdi_raw, label="Ecuador (IDH Ajustado por Presiones Planetarias - PHDI)", color="#16A085", marker="s", linewidth=2.5)
plt.title("Comparativo Internacional: IDH Ajustado por Presiones Planetarias (PHDI) (2000-2020)", fontsize=12, fontweight="bold", pad=15)
plt.xlabel("Año", fontsize=10, fontweight="semibold")
plt.ylabel("PHDI", fontsize=10, fontweight="semibold")
plt.xticks(years, rotation=45)
plt.ylim(0.50, 1.00)
plt.legend(loc="lower right", frameon=True, facecolor="white")
plt.tight_layout()
plt.savefig(f"{base_dir}/assets/phdi_comparison.png", dpi=300)
plt.close()

# Chart 4: GII Comparison (Desigualdad de Género)
plt.figure(figsize=(10, 6))
plt.plot(years, che_gii_raw, label="Suiza (Índice de Desigualdad de Género - IDG)", color="#C0392B", marker="o", linewidth=2.5)
plt.plot(years, ecu_gii_raw, label="Ecuador (Índice de Desigualdad de Género - IDG)", color="#D35400", marker="s", linewidth=2.5)
plt.title("Comparativo Internacional: Índice de Desigualdad de Género (IDG / GII) (2000-2020)", fontsize=12, fontweight="bold", pad=15)
plt.xlabel("Año", fontsize=10, fontweight="semibold")
plt.ylabel("IDG (Menos es Mejor)", fontsize=10, fontweight="semibold")
plt.xticks(years, rotation=45)
plt.ylim(0.00, 0.60)
plt.legend(loc="upper right", frameon=True, facecolor="white")
plt.tight_layout()
plt.savefig(f"{base_dir}/assets/gii_comparison.png", dpi=300)
plt.close()

log_message("Gráficos de los 4 indicadores guardados exitosamente en assets/.")

# 6. Generate the Premium Styled Excel File using openpyxl
log_message("Construyendo libro de Excel premium con hojas y fórmulas vivas...")
wb = Workbook()

# Style constants
navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
slate_fill = PatternFill(start_color="4F5D73", end_color="4F5D73", fill_type="solid")
light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
soft_blue_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")

white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
bold_font = Font(name="Calibri", size=11, bold=True)
regular_font = Font(name="Calibri", size=11)
title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center")
right_alignment = Alignment(horizontal="right", vertical="center")

thin_side = Side(border_style="thin", color="D3D3D3")
double_side = Side(border_style="double", color="000000")
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_total = Border(top=thin_side, bottom=double_side)

def populate_country_sheet(ws, country_name, health_raw, expected_raw, mean_raw, gni_raw, unemp_rates, official_hdi, ihdi_raw, phdi_raw, gii_raw):
    ws.views.sheetView[0].showGridLines = True
    
    # Title blocks
    ws.merge_cells("A1:Q1")
    ws["A1"] = f"Indicadores de Desarrollo Humano (IDH, IHDI, PHDI, IDG) - {country_name} (2000-2020)"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_alignment
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:Q2")
    ws["A2"] = "Materia: Desarrollo Económico | Taller: Análisis Comparativo Internacional | Estudiante: Erick Fabricio Condoy Seraquive"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = left_alignment
    
    headers = [
        "Año", 
        "Esp. Vida (Años)", 
        "Índice Salud (I_Salud)", 
        "Años Esp. Esc.", 
        "Índice Años Esp.", 
        "Años Prom. Esc.", 
        "Índice Años Prom.", 
        "Índice Educación (I_Edu)", 
        "INB per cápita PPA (USD)", 
        "Índice Ingresos (I_Ing)", 
        "Tasa Desempleo (%)", 
        "Índice Empleo (I_Emp)", 
        "IDH Oficial (Calculado)", 
        "IDH Oficial (PNUD Raw)", 
        "IHDI (Ajust. Desigualdad)",
        "PHDI (Ajust. Planeta)",
        "IDG / GII (Desigualdad Género)"
    ]
    
    ws.row_dimensions[4].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = center_alignment
        cell.border = border_all
        
    for i in range(len(years)):
        row_num = 5 + i
        ws.row_dimensions[row_num].height = 20
        
        # Populate values
        ws.cell(row=row_num, column=1, value=str(int(years[i]))).alignment = center_alignment
        ws.cell(row=row_num, column=2, value=float(health_raw[i])).number_format = "0.000"
        
        # Salud: =(B{row_num}-20)/65
        ws.cell(row=row_num, column=3, value=f"=(B{row_num}-20)/65").number_format = "0.000"
        
        ws.cell(row=row_num, column=4, value=float(expected_raw[i])).number_format = "0.000"
        
        # Años esp: =D{row_num}/18
        ws.cell(row=row_num, column=5, value=f"=D{row_num}/18").number_format = "0.000"
        
        ws.cell(row=row_num, column=6, value=float(mean_raw[i])).number_format = "0.000"
        
        # Años prom: =F{row_num}/15
        ws.cell(row=row_num, column=7, value=f"=F{row_num}/15").number_format = "0.000"
        
        # Edu: =(E{row_num}+G{row_num})/2
        ws.cell(row=row_num, column=8, value=f"=(E{row_num}+G{row_num})/2").number_format = "0.000"
        
        ws.cell(row=row_num, column=9, value=float(gni_raw[i])).number_format = "$#,##0.00"
        
        # Ingreso capped at 75000: =(LN(MIN(I{row_num},75000))-LN(100))/(LN(75000)-LN(100))
        ws.cell(row=row_num, column=10, value=f"=(LN(MIN(I{row_num},75000))-LN(100))/(LN(75000)-LN(100))").number_format = "0.000"
        
        ws.cell(row=row_num, column=11, value=float(unemp_rates[i]) / 100.0).number_format = "0.00%"
        
        # Empleo: =(0.15-K{row_num})/0.15
        ws.cell(row=row_num, column=12, value=f"=(0.15-K{row_num})/0.15").number_format = "0.000"
        
        # IDH Calculado: =(C{row_num}*H{row_num}*J{row_num})^(1/3)
        ws.cell(row=row_num, column=13, value=f"=(C{row_num}*H{row_num}*J{row_num})^(1/3)").number_format = "0.000"
        
        # Raw indicators from CSV
        ws.cell(row=row_num, column=14, value=float(official_hdi[i]) if not np.isnan(official_hdi[i]) else "").number_format = "0.000"
        ws.cell(row=row_num, column=15, value=float(ihdi_raw[i]) if not np.isnan(ihdi_raw[i]) else "").number_format = "0.000"
        ws.cell(row=row_num, column=16, value=float(phdi_raw[i]) if not np.isnan(phdi_raw[i]) else "").number_format = "0.000"
        ws.cell(row=row_num, column=17, value=float(gii_raw[i]) if not np.isnan(gii_raw[i]) else "").number_format = "0.000"
        
        for col_idx in range(1, 18):
            c = ws.cell(row=row_num, column=col_idx)
            c.font = regular_font
            c.border = border_all
            if col_idx > 1:
                c.alignment = right_alignment
            if i % 2 == 1:
                c.fill = light_gray_fill
                
    # Average row
    avg_row = 5 + len(years)
    ws.row_dimensions[avg_row].height = 22
    ws.cell(row=avg_row, column=1, value="Promedio").font = bold_font
    ws.cell(row=avg_row, column=1).alignment = center_alignment
    ws.cell(row=avg_row, column=1).border = border_total
    
    for col_idx in range(2, 18):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=avg_row, column=col_idx, value=f"=AVERAGE({col_letter}5:{col_letter}{avg_row-1})")
        cell.font = bold_font
        cell.alignment = right_alignment
        cell.border = border_total
        
        if col_idx in [2, 4, 6]:
            cell.number_format = "0.00"
        elif col_idx == 9:
            cell.number_format = "$#,##0.00"
        elif col_idx == 11:
            cell.number_format = "0.00%"
        else:
            cell.number_format = "0.000"
            
    # Adjust widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

# Create Ecuador calculations sheet
ws_ecu = wb.active
ws_ecu.title = "Ecuador_Calculos"
populate_country_sheet(ws_ecu, "Ecuador", ecu_health_raw, ecu_expected_raw, ecu_mean_raw, ecu_gni_raw, ecu_unemployment, ecu_official_hdi, ecu_ihdi_raw, ecu_phdi_raw, ecu_gii_raw)

# Create Switzerland calculations sheet
ws_che = wb.create_sheet("Suiza_Calculos")
populate_country_sheet(ws_che, "Suiza", che_health_raw, che_expected_raw, che_mean_raw, che_gni_raw, che_unemployment, che_official_hdi, che_ihdi_raw, che_phdi_raw, che_gii_raw)


# --- SHEET 3: COMPARATIVO ---
ws_comp = wb.create_sheet("Comparativo_Indicadores")
ws_comp.views.sheetView[0].showGridLines = True

ws_comp.merge_cells("A1:I1")
ws_comp["A1"] = "Comparativo Lado a Lado de los 4 Indicadores: Ecuador vs. Suiza (2000-2020)"
ws_comp["A1"].font = title_font
ws_comp["A1"].alignment = left_alignment
ws_comp.row_dimensions[1].height = 30

ws_comp.merge_cells("A2:I2")
ws_comp["A2"] = "Análisis comparativo de los indicadores oficiales de PNUD extraídos fielmente de las bases raw."
ws_comp["A2"].font = Font(name="Calibri", size=10, italic=True)
ws_comp["A2"].alignment = left_alignment

comp_headers = [
    "Año",
    "Ecuador IDH",
    "Suiza IDH",
    "Ecuador IHDI (Desigualdad)",
    "Suiza IHDI (Desigualdad)",
    "Ecuador PHDI (Planeta)",
    "Suiza PHDI (Planeta)",
    "Ecuador IDG / GII (Género)",
    "Suiza IDG / GII (Género)"
]

ws_comp.row_dimensions[4].height = 25
for col_idx, header in enumerate(comp_headers, 1):
    cell = ws_comp.cell(row=4, column=col_idx, value=header)
    cell.fill = slate_fill
    cell.font = white_bold_font
    cell.alignment = center_alignment
    cell.border = border_all

for i in range(len(years)):
    row_num = 5 + i
    ws_comp.row_dimensions[row_num].height = 20
    
    # Year
    ws_comp.cell(row=row_num, column=1, value=str(int(years[i]))).alignment = center_alignment
    
    # IDH
    ws_comp.cell(row=row_num, column=2, value=f"=Ecuador_Calculos!N{row_num}").number_format = "0.000"
    ws_comp.cell(row=row_num, column=3, value=f"=Suiza_Calculos!N{row_num}").number_format = "0.000"
    
    # IHDI
    ws_comp.cell(row=row_num, column=4, value=f"=Ecuador_Calculos!O{row_num}").number_format = "0.000"
    ws_comp.cell(row=row_num, column=5, value=f"=Suiza_Calculos!O{row_num}").number_format = "0.000"
    
    # PHDI
    ws_comp.cell(row=row_num, column=6, value=f"=Ecuador_Calculos!P{row_num}").number_format = "0.000"
    ws_comp.cell(row=row_num, column=7, value=f"=Suiza_Calculos!P{row_num}").number_format = "0.000"
    
    # GII
    ws_comp.cell(row=row_num, column=8, value=f"=Ecuador_Calculos!Q{row_num}").number_format = "0.000"
    ws_comp.cell(row=row_num, column=9, value=f"=Suiza_Calculos!Q{row_num}").number_format = "0.000"
    
    for col_idx in range(1, 10):
        c = ws_comp.cell(row=row_num, column=col_idx)
        c.font = regular_font
        c.border = border_all
        if col_idx > 1:
            c.alignment = right_alignment
        if i % 2 == 1:
            c.fill = light_gray_fill

# Average row for Comparativo
avg_row = 5 + len(years)
ws_comp.row_dimensions[avg_row].height = 22
ws_comp.cell(row=avg_row, column=1, value="Promedio").font = bold_font
ws_comp.cell(row=avg_row, column=1).alignment = center_alignment
ws_comp.cell(row=avg_row, column=1).border = border_total

for col_idx in range(2, 10):
    col_letter = get_column_letter(col_idx)
    cell = ws_comp.cell(row=avg_row, column=col_idx, value=f"=AVERAGE({col_letter}5:{col_letter}{avg_row-1})")
    cell.font = bold_font
    cell.alignment = right_alignment
    cell.border = border_total
    cell.number_format = "0.000"

for col in ws_comp.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws_comp.column_dimensions[col_letter].width = max(max_len + 4, 18)


# --- NATIVE LINE CHARTS SHEETS ---

def add_native_chart_sheet(sheet_title, data_ws, chart_title, min_col_data, max_col_data, chart_y_title, series_names_colors):
    ws_fig = wb.create_sheet(title=sheet_title)
    ws_fig.views.sheetView[0].showGridLines = True
    
    ws_fig.row_dimensions[1].height = 25
    ws_fig.cell(row=1, column=1, value=chart_title).font = title_font
    
    chart = LineChart()
    chart.title = chart_title
    chart.style = 13
    chart.y_axis.title = chart_y_title
    chart.x_axis.title = "Año"
    chart.width = 18
    chart.height = 12
    
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    
    max_row_num = 4 + len(years)
    data_ref = Reference(data_ws, min_col=min_col_data, min_row=4, max_col=max_col_data, max_row=max_row_num)
    cats_ref = Reference(data_ws, min_col=1, min_row=5, max_row=max_row_num)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    for idx, (name, color) in enumerate(series_names_colors):
        s = chart.series[idx]
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = 38100  # 3 pt
        s.marker.symbol = "circle" if idx == 0 else "square"
        s.marker.size = 6
        s.marker.graphicalProperties.solidFill = color
        s.marker.graphicalProperties.line.solidFill = color
        s.smooth = True
        
    ws_fig.add_chart(chart, "A3")

add_native_chart_sheet("Grafico_IDH", ws_comp, "IDH Oficial: Ecuador vs Suiza", 2, 3, "Valor del Índice", [("Ecuador IDH", "3498DB"), ("Suiza IDH", "2C3E50")])
add_native_chart_sheet("Grafico_IHDI", ws_comp, "IDH Ajustado por Desigualdad (IHDI): Ecuador vs Suiza", 4, 5, "Valor del Índice", [("Ecuador IHDI", "F1C40F"), ("Suiza IHDI", "27AE60")])
add_native_chart_sheet("Grafico_PHDI", ws_comp, "IDH Ajustado por Presiones Planetarias (PHDI)", 6, 7, "Valor del Índice", [("Ecuador PHDI", "16A085"), ("Suiza PHDI", "8E44AD")])
add_native_chart_sheet("Grafico_IDG_GII", ws_comp, "Índice de Desigualdad de Género (IDG / GII)", 8, 9, "Valor (Menos es Mejor)", [("Ecuador IDG", "D35400"), ("Suiza IDG", "C0392B")])


# --- SHEET 8: DICTIONARY ---
ws_dict = wb.create_sheet(title="Metadatos_Diccionario")
ws_dict.views.sheetView[0].showGridLines = True

ws_dict.merge_cells("A1:F1")
ws_dict["A1"] = "Diccionario de Metadatos de Desarrollo Humano Oficial"
ws_dict["A1"].font = title_font
ws_dict["A1"].alignment = left_alignment
ws_dict.row_dimensions[1].height = 30

ws_dict.merge_cells("A2:F2")
ws_dict["A2"] = "Especificación de variables de los 4 indicadores oficiales del PNUD y su enlace de descarga."
ws_dict["A2"].font = Font(name="Calibri", size=10, italic=True)
ws_dict["A2"].alignment = left_alignment

dict_headers = ["Variable", "Definición Conceptual", "Fórmula / Dimensiones", "Unidad de Medida", "Fuente de Origen", "Enlace de Descarga"]
ws_dict.row_dimensions[4].height = 25
for col_idx, header in enumerate(dict_headers, 1):
    cell = ws_dict.cell(row=4, column=col_idx, value=header)
    cell.fill = navy_fill
    cell.font = white_bold_font
    cell.alignment = center_alignment
    cell.border = border_all

dict_data = [
    {
        "variable": "IDH (Índice de Desarrollo Humano)",
        "definition": "Medida sintética que evalúa el progreso medio de un país en tres dimensiones básicas: salud, educación y nivel de vida.",
        "formula": "=(I_Salud * I_Edu * I_Ing)^(1/3)",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Human Development Reports)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "IHDI (IDH Ajustado por la Desigualdad)",
        "definition": "Evalúa el nivel de desarrollo humano ajustando el valor de cada dimensión en función de la desigualdad en su distribución entre la población.",
        "formula": "=IDH * (1 - Inequidad_Promedio)",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Atkinson's Inequality Index)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "PHDI (IDH Ajustado por Presiones Planetarias)",
        "definition": "Ajusta el IDH oficial multiplicándolo por un factor de presión que considera las emisiones per cápita de CO2 y la huella material.",
        "formula": "=IDH * S_Factor_Presión_Ambiental",
        "unit": "Índice (0 a 1)",
        "source": "PNUD (Planetary Pressures)",
        "link": "https://hdr.undp.org/data-center"
    },
    {
        "variable": "IDG / GII (Índice de Desigualdad de Género)",
        "definition": "Mide las desventajas de las mujeres en tres dimensiones básicas: salud reproductiva, empoderamiento y mercado laboral. Cero significa igualdad perfecta.",
        "formula": "Fórmula combinada de medias geométricas por género (GII)",
        "unit": "Índice (0 a 1), donde menor es mejor",
        "source": "PNUD (Gender Gaps)",
        "link": "https://hdr.undp.org/data-center"
    }
]

for row_idx, data in enumerate(dict_data, 5):
    ws_dict.row_dimensions[row_idx].height = 45
    
    ws_dict.cell(row=row_idx, column=1, value=data["variable"]).font = bold_font
    ws_dict.cell(row=row_idx, column=1).alignment = left_alignment
    ws_dict.cell(row=row_idx, column=1).border = border_all
    
    c_def = ws_dict.cell(row=row_idx, column=2, value=data["definition"])
    c_def.font = regular_font
    c_def.alignment = left_alignment
    c_def.border = border_all
    c_def.alignment = Alignment(wrap_text=True, vertical="center")
    
    c_form = ws_dict.cell(row=row_idx, column=3, value=data["formula"])
    c_form.font = regular_font
    c_form.alignment = center_alignment
    c_form.border = border_all
    c_form.alignment = Alignment(wrap_text=True, vertical="center")
    
    ws_dict.cell(row=row_idx, column=4, value=data["unit"]).font = regular_font
    ws_dict.cell(row=row_idx, column=4).alignment = left_alignment
    ws_dict.cell(row=row_idx, column=4).border = border_all
    
    c_src = ws_dict.cell(row=row_idx, column=5, value=data["source"])
    c_src.font = regular_font
    c_src.alignment = left_alignment
    c_src.border = border_all
    c_src.alignment = Alignment(wrap_text=True, vertical="center")
    
    cell_link = ws_dict.cell(row=row_idx, column=6, value=data["link"])
    cell_link.border = border_all
    if data["link"].startswith("http"):
        cell_link.hyperlink = data["link"]
        cell_link.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        cell_link.alignment = left_alignment
    else:
        cell_link.font = regular_font
        cell_link.alignment = left_alignment
        
    if row_idx % 2 == 1:
        for col_idx in range(1, 7):
            ws_dict.cell(row=row_idx, column=col_idx).fill = light_gray_fill

# Adjust column widths for Dictionary
ws_dict.column_dimensions["A"].width = 30
ws_dict.column_dimensions["B"].width = 45
ws_dict.column_dimensions["C"].width = 30
ws_dict.column_dimensions["D"].width = 20
ws_dict.column_dimensions["E"].width = 30
ws_dict.column_dimensions["F"].width = 35

excel_out_path = f"{base_dir}/data/processed/comparative_hdi_expanded_20y.xlsx"
wb.save(excel_out_path)
log_message(f"Libro de Excel maestro guardado exitosamente en: {excel_out_path}")
log_message("=== CÁLCULO Y CONSTRUCCIÓN COMPLETADOS CON ÉXITO ===")
